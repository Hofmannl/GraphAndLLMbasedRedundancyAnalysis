import os
import pandas as pd
from multiprocessing import Process, Queue
from time import sleep
from collections.abc import Callable
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter as utils_get_column_letter
from dotenv import load_dotenv
from .load_data import load_datasets_with_out_annotations as loading

load_dotenv()


def check_if_sheet_exists_and_delete(q: Queue) -> None:
    delete_flag: bool = False
    path_to_file: str = str(q.get())
    file_name: str = str(q.get())
    sheet_name: str = str(q.get())
    full_path: str = os.path.join(path_to_file, file_name)
    try:
        wb = load_workbook(full_path)
        if sheet_name in wb.sheetnames and len(wb.sheetnames) == 1:
            delete_flag = True
        else:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            wb.save(full_path)
    except FileNotFoundError:
        print(f"File not found: {full_path}")
    except PermissionError:
        print(f"Permission denied while accessing {full_path}")
    except Exception as e:
        print(f"An unexpected error occurred while processing {full_path}: {e}")
    finally:
        wb.close()
        q.put(delete_flag)


def save_to_excel(
    local_data: pd.DataFrame,
    formatter: Callable[[Workbook, str], None] = None,
    sheet_name: str = "Sheet",
    name_xlsx: str = os.getenv("OUTPUT_EXCEL_NAME_WITHOUT_ANNOTATIONS"),
):
    mode: str = "w"
    if os.path.exists(name_xlsx):
        q: Queue = Queue()
        q.put(os.getcwd())
        q.put(name_xlsx)
        q.put(sheet_name)
        process = Process(
            target=check_if_sheet_exists_and_delete,
            args=(q,),
        )
        process.start()
        process.join()
        while process.is_alive():
            sleep(1)
        process.close()
        if not q.empty():
            if bool(q.get()):
                full_file_path: str = os.path.join(os.getcwd(), name_xlsx)
                os.remove(full_file_path)
            else:
                mode = "r+"
    with pd.ExcelWriter(name_xlsx, mode=mode) as writer:
        local_data.to_excel(writer, index=False, sheet_name=sheet_name)
    if formatter:
        wb = load_workbook(name_xlsx)
        formatter(wb, sheet_name)
        wb.save(name_xlsx)
        wb.close()


def load_datasets_add_line_counter() -> dict[str, list]:
    datasets: dict[str, list] = loading()
    for key in datasets.keys():
        current_dataset = datasets[key]
        for i, item in enumerate(current_dataset, start=1):
            item = {"linecounter": i, **item}
            current_dataset[i - 1] = item
    return datasets


def prepaire_excel_data(
    sheet_name: str,
    suffix: str = "\\src_gpt_approach",
    file_name: str = "Evaluation_Graph.xlsx",
) -> pd:
    current_directory = os.getcwd()
    suffix = suffix
    directory_excel = (
        current_directory[: -len(suffix)]
        if current_directory.endswith(suffix)
        else current_directory
    )
    directory_excel += f"\\Datasets\\{file_name}"
    excel_data = pd.read_excel(
        directory_excel,
        usecols=lambda column: "Item" not in column,
        skiprows=range(1),
        sheet_name=sheet_name,
    )
    # excel_data = excel_data.drop(
    #     columns=[
    #         "Unnamed: 0",
    #         "Unnamed: 12",
    #         "total",
    #         "main",
    #         "benefit",
    #         "total.1",
    #         "main.1",
    #         "benefit.1",
    #     ],
    #     axis=1,
    #     errors="ignore",
    # )
    excel_data = excel_data.rename(
        columns={
            "Main Part \nPartial": "Main Part Partial",
            "Main Part \nFull": "Main Part Full",
            "Benefit\nPartial": "Benefit Partial",
            "Benefit\nFull": "Benefit Full",
        }
    )
    excel_data.index += 1
    excel_data = excel_data.fillna("Empty")

    ## Map line number to user story number
    # excel_data.insert(loc=4, column="Corresponding USID 1", value=0)
    # excel_data.insert(loc=5, column="Corresponding USID 2", value=0)
    # datasets = load_datasets_add_line_counter()
    # for idx in range(len(excel_data)):
    #     redundant_pairs = excel_data.iat[idx, 1]
    #     parts = redundant_pairs.split("_")
    #     first_number = int(parts[2])
    #     second_number = int(parts[-1])
    #     project_number = f"#{excel_data.iat[idx, 0]}#".upper()
    #     project_data = datasets[project_number]
    #     for item in project_data:
    #         if item["linecounter"] == first_number:
    #             excel_data.iat[idx, 4] = item["id"]
    #         if item["linecounter"] == second_number:
    #             excel_data.iat[idx, 5] = item["id"]
    return excel_data


def formatter_ignored_items(wb: Workbook, sheet_name: str):
    ws = wb[sheet_name]
    header_font = Font(size=14, bold=True)
    for cell in ws["1:1"]:
        cell.font = header_font

    ADDITIONAL_LENGTH: int = 0
    ADJUSTED_WIDTH: int = 0
    MAX_LEN: int = 0
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            MAX_LEN = len(str(cell.value))
            ADDITIONAL_LENGTH = MAX_LEN + 2
            ADJUSTED_WIDTH = 0

            if cell.col_idx == 1:
                ADJUSTED_WIDTH = ADDITIONAL_LENGTH * 10
            if cell.col_idx == 2:
                ADJUSTED_WIDTH = ADDITIONAL_LENGTH * 3
            else:
                ADJUSTED_WIDTH = ADDITIONAL_LENGTH * 1.5
            ws.column_dimensions[utils_get_column_letter(cell.column)].width = (
                ADJUSTED_WIDTH
            )

    alignment = Alignment(vertical="center", horizontal="left")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    num_columns = ws.max_column
    header_range = f"A1:{utils_get_column_letter(num_columns)}1"
    ws.auto_filter.ref = header_range
    ws.freeze_panes = ws["A2"]

    wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
